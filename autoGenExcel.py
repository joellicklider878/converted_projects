import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage
import matplotlib.pyplot as plt
from datetime import datetime
import calendar

wb = openpyxl.Workbook()
sheet = wb.active

sheet["A1"].value = "Month"
sheet["B1"].value = "Number of Days"
for col in range(1, 3):
    col_letter = get_column_letter(col)
    sheet[f"{col_letter}1"].font = openpyxl.styles.Font(size=12, bold=True)
    sheet.column_dimensions[col_letter].width = 20
sheet.row_dimensions[1].height = 20

for i in range(1, 13):
    days = calendar.monthrange(datetime.now().year, i)[1]
    month_name = calendar.month_name[i]
    sheet[f"A{i + 1}"].value = month_name
    sheet[f"B{i + 1}"].value = days

months = [sheet[f"A{i}"].value for i in range(2, 14)]
days_in_month = [sheet[f"B{i}"].value for i in range(2, 14)]

plt.figure(figsize=(10, 6))
plt.bar(months, days_in_month)
plt.xlabel('Month')
plt.ylabel('Number of Days')
plt.title('Number of Days in Each Month')
plt.xticks(rotation=45)
plt.tight_layout()

chart_image_path = "chart.png"
plt.savefig(chart_image_path)

chart_img = ExcelImage(chart_image_path)
sheet.add_image(chart_img, "D1")

wb.save("DriveExcelForm.xlsx")
